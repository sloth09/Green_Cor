"""
Export optimization results to Excel format.
Creates multi-sheet workbook with scenario summary and yearly results.
"""

from pathlib import Path
from typing import Dict, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    """Export MILP results to Excel workbook."""

    def __init__(self, config: Dict):
        """
        Initialize Excel exporter.

        Args:
            config: Configuration dictionary
        """
        self.config = config
        self.case_name = config.get("case_name", "Unknown")
        self.case_id = config.get("case_id", "unknown")

    def export_results(
        self,
        scenario_df: pd.DataFrame,
        yearly_df: pd.DataFrame,
        output_path: Path = None
    ) -> Path:
        """
        Export results to Excel file.

        Args:
            scenario_df: Scenario summary DataFrame
            yearly_df: Yearly results DataFrame
            output_path: Output directory (default: results/)

        Returns:
            Path to created Excel file
        """
        if output_path is None:
            output_path = Path("results")

        output_path.mkdir(parents=True, exist_ok=True)

        filename = f"MILP_results_{self.case_id}.xlsx"
        filepath = output_path / filename

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add sheets
        self._add_summary_sheet(wb, scenario_df)
        self._add_time_breakdown_sheet(wb, scenario_df)
        self._add_yearly_sheet(wb, yearly_df)
        self._add_config_sheet(wb)

        # Save
        wb.save(filepath)
        print(f"Excel export completed: {filepath}")

        return filepath

    def _add_summary_sheet(self, wb: Workbook, scenario_df: pd.DataFrame) -> None:
        """
        Add scenario summary sheet.

        Args:
            wb: Workbook object
            scenario_df: Scenario DataFrame
        """
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = f"Scenario Summary - {self.case_name}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # Sort by NPC
        df_sorted = scenario_df.sort_values("NPC_Total_USDm").copy()

        # Write headers
        headers = df_sorted.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, row in enumerate(dataframe_to_rows(df_sorted, index=False, header=False), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format numeric columns
                if col_idx in [3, 4, 5] or col_idx > 5:  # NPC columns
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="center")

                # Highlight top 3
                if row_idx <= 6:  # Top 3 scenarios (rows 4-6)
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        for col in ["C", "D", "E", "F", "G", "H", "I", "J"]:
            ws.column_dimensions[col].width = 18

        # Add statistics
        stats_row = len(scenario_df) + 5

        ws.cell(row=stats_row, column=1).value = "Statistics"
        ws.cell(row=stats_row, column=1).font = Font(bold=True)

        ws.cell(row=stats_row + 1, column=1).value = "Total Scenarios:"
        ws.cell(row=stats_row + 1, column=2).value = len(scenario_df)

        ws.cell(row=stats_row + 2, column=1).value = "Best NPC (M USD):"
        ws.cell(row=stats_row + 2, column=2).value = df_sorted["NPC_Total_USDm"].min()
        ws.cell(row=stats_row + 2, column=2).number_format = "0.00"

        ws.cell(row=stats_row + 3, column=1).value = "Worst NPC (M USD):"
        ws.cell(row=stats_row + 3, column=2).value = df_sorted["NPC_Total_USDm"].max()
        ws.cell(row=stats_row + 3, column=2).number_format = "0.00"

        ws.cell(row=stats_row + 4, column=1).value = "Average NPC (M USD):"
        ws.cell(row=stats_row + 4, column=2).value = df_sorted["NPC_Total_USDm"].mean()
        ws.cell(row=stats_row + 4, column=2).number_format = "0.00"

    def _add_time_breakdown_sheet(self, wb: Workbook, scenario_df: pd.DataFrame) -> None:
        """
        Add time breakdown sheet for optimal scenario.

        Args:
            wb: Workbook object
            scenario_df: Scenario DataFrame
        """
        ws = wb.create_sheet("Time Breakdown", 1)

        # Find optimal scenario (minimum NPC)
        optimal_idx = scenario_df["NPC_Total_USDm"].idxmin()
        optimal = scenario_df.loc[optimal_idx]

        # Title
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = f"【최적 시나리오 시간 분석】"
        title_cell.font = Font(size=12, bold=True)
        ws.row_dimensions[1].height = 20

        # Case and scenario info
        row = 3
        ws.cell(row=row, column=1).value = "Case"
        ws.cell(row=row, column=2).value = self.case_name
        ws.cell(row=row + 1, column=1).value = "Shuttle Size (m³)"
        ws.cell(row=row + 1, column=2).value = int(optimal["Shuttle_Size_cbm"])
        ws.cell(row=row + 2, column=1).value = "Pump Size (m³/h)"
        ws.cell(row=row + 2, column=2).value = int(optimal["Pump_Size_m3ph"])
        ws.cell(row=row + 3, column=1).value = "NPC (M USD)"
        ws.cell(row=row + 3, column=2).value = optimal["NPC_Total_USDm"]
        ws.cell(row=row + 3, column=2).number_format = "0.00"
        ws.cell(row=row + 4, column=1).value = "LCOA (USD/ton)"
        ws.cell(row=row + 4, column=2).value = optimal.get("LCOAmmonia_USD_per_ton", 0)
        ws.cell(row=row + 4, column=2).number_format = "0.00"

        # Time breakdown table
        row = 9
        ws.merge_cells(f"A{row}:C{row}")
        header_cell = ws[f"A{row}"]
        header_cell.value = "1회 왕복 운항 시간 분해 (Time Breakdown)"
        header_cell.font = Font(bold=True, size=11)

        # Table header
        row = 10
        headers = ["시간 구성 요소", "시간 (h)", "비율 (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Time components
        basic_cycle = optimal.get("Basic_Cycle_Duration_hr", 0)
        components = [
            ("육상 적재", optimal.get("Shore_Loading_hr", 0)),
            ("편도 항해", optimal.get("Travel_Outbound_hr", 0)),
            ("호스 연결", optimal.get("Setup_Inbound_hr", 0)),
            ("펌핑", optimal.get("Pumping_Per_Vessel_hr", 0)),
            ("호스 분리", optimal.get("Setup_Outbound_hr", 0)),
            ("복귀 항해", optimal.get("Travel_Return_hr", 0)),
        ]

        row = 11
        total_hours = 0
        for name, hours in components:
            ws.cell(row=row, column=1).value = name
            ws.cell(row=row, column=2).value = round(hours, 2)
            ws.cell(row=row, column=2).number_format = "0.00"
            if basic_cycle > 0:
                percentage = (hours / basic_cycle) * 100
            else:
                percentage = 0
            ws.cell(row=row, column=3).value = round(percentage, 1)
            ws.cell(row=row, column=3).number_format = "0.0\%"
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
            total_hours += hours
            row += 1

        # Total
        ws.cell(row=row, column=1).value = "【총 사이클 시간】"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = optimal["Cycle_Duration_hr"]
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = "0.00"
        ws.cell(row=row, column=3).value = "100%"
        ws.cell(row=row, column=3).font = Font(bold=True)

        # Operating metrics
        row += 2
        ws.cell(row=row, column=1).value = "기본 사이클 (육상 제외)"
        ws.cell(row=row, column=2).value = optimal.get("Basic_Cycle_Duration_hr", 0)
        ws.cell(row=row, column=2).number_format = "0.00"

        row += 2
        ws.cell(row=row, column=1).value = "연간 운영 지표"
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)

        row += 1
        metrics = [
            ("연간 최대 항차", optimal["Annual_Cycles_Max"], "회"),
            ("연간 공급 용량", optimal["Annual_Supply_m3"], "m³"),
            ("시간 활용도", optimal["Time_Utilization_Ratio_percent"], "%"),
            ("선박당 일정", 365 / optimal["Annual_Cycles_Max"] if optimal["Annual_Cycles_Max"] > 0 else 0, "일"),
        ]

        for name, value, unit in metrics:
            ws.cell(row=row, column=1).value = name
            ws.cell(row=row, column=2).value = round(value, 2) if isinstance(value, float) else value
            ws.cell(row=row, column=3).value = unit
            if isinstance(value, float):
                ws.cell(row=row, column=2).number_format = "0.00"
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12

    def _add_yearly_sheet(self, wb: Workbook, yearly_df: pd.DataFrame) -> None:
        """
        Add yearly results sheet.

        Args:
            wb: Workbook object
            yearly_df: Yearly results DataFrame
        """
        ws = wb.create_sheet("Yearly Results", 1)

        # Title
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"Yearly Results - {self.case_name}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # Write headers
        headers = yearly_df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, row in enumerate(dataframe_to_rows(yearly_df, index=False, header=False), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format numeric columns
                if col_idx > 3:
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="center")

        # Freeze panes
        ws.freeze_panes = "A4"

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 16

    def _add_config_sheet(self, wb: Workbook) -> None:
        """
        Add configuration sheet.

        Args:
            wb: Workbook object
        """
        ws = wb.create_sheet("Configuration", 2)

        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = f"Configuration - {self.case_name}"
        title_cell.font = Font(size=12, bold=True)

        # Key configuration parameters
        row = 3
        ws.cell(row=row, column=1).value = "Parameter"
        ws.cell(row=row, column=2).value = "Value"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)

        params = [
            ("Case Name", self.config.get("case_name", "Unknown")),
            ("Case ID", self.config.get("case_id", "unknown")),
            ("Time Period", f"{self.config['time_period']['start_year']}-{self.config['time_period']['end_year']}"),
            ("Discount Rate", f"{self.config['economy']['discount_rate']*100}%"),
            ("Fuel Price (USD/ton)", self.config['economy']['fuel_price_usd_per_ton']),
            ("Bunker Volume per Call (m3)", self.config['bunkering']['bunker_volume_per_call_m3']),
            ("Travel Time (hours)", self.config['operations']['travel_time_hours']),
            ("Max Annual Hours", self.config['operations']['max_annual_hours_per_vessel']),
            ("Tank Storage Enabled", self.config['tank_storage']['enabled']),
            ("Tank Size (tons)", self.config['tank_storage']['size_tons']),
            ("Start Vessels", self.config['shipping']['start_vessels']),
            ("End Vessels", self.config['shipping']['end_vessels']),
        ]

        for idx, (param, value) in enumerate(params, row + 1):
            ws.cell(row=idx, column=1).value = param
            ws.cell(row=idx, column=2).value = value
            ws.cell(row=idx, column=2).alignment = Alignment(horizontal="right")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
